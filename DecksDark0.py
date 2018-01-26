
import openpyxl
import sys
from geopy.distance import vincenty


def distanceCalc(x, y):
    return vincenty(x, y).miles


def diff(m, n):
    if (m == None or n == None):
        return 0
    return (int(n) - int(m))
def foreachmla(xyz,res,m,sprintcell,location1,maxrow,mlaVB):
        for xyrow in range(2, maxrow + 1):
            print m
            if (str(xyz['E' + str(xyrow)].value)) == m:
                muniq = xyz['A' + str(xyrow)].value
                mlat = xyz['B' + str(xyrow)].value
                mlng = xyz['C' + str(xyrow)].value
                location3 = (mlat,mlng)
                dis = distanceCalc(location1,location3)
                mlaVB.append({"mind":dis,"mlauniq":muniq})
        my_list = sorted(mlaVB, key=lambda k: k['mind'])
        return my_list[0]
def comparestuff(xyz,res,sprintcell,location1,rad, maxrow, l, v):
    for xyrow in range(2, maxrow + 1):
        xyuniq = xyz['A' + str(xyrow)].value
        xylat = xyz['B' + str(xyrow)].value
        xylng = xyz['C' + str(xyrow)].value
        xyrad = xyz['D' + str(xyrow)].value
        xymla = xyz['E' + str(xyrow)].value
        location2 = (xylat,xylng)
        dis = distanceCalc(location1,location2)
        diffh = diff(rad,xyrad)
        l.append(dis)
        v.append({"cellsiteXYZ": xyuniq, "SprintCellsite": sprintcell, "location1": location1, "location2": location2,
                  "distance": dis, "heightdiff": diffh,"xyrad":xyrad,"xymla":xymla })
    # print min(l)
    minIndex = l.index(min(l))
    return v[minIndex]


def doExcelProc(wb,hq,dq):
    print (wb.sheetnames)

    # get sheets
    oc = wb['OurCompany']
    xyz = wb['XYZcompany']
    # get sheet max and min rows and cols
    oc_maxrow = oc.max_row
    oc_maxcol = oc.max_column
    xyz_maxrow = xyz.max_row
    xyz_maxcol = xyz.max_column

    # make new Sheet
    wb.create_sheet(index=3, title='Results')
    res = wb['Results']
    res['A1'].value = "SprintCell"
    res['B1'].value = "Sprint Location"
    res['C1'].value = "XYZCell"
    res['D1'].value = "XYZCell location"
    res['E1'].value = "Minimum Distance"
    res['F1'].value = "RAD center of xyzcompany"
    res['G1'].value = "Height of sprint tower"
    res['H1'].value = "RAD center difference"
    res['I1'].value = "Sites with height difference of"+ str(hq) + "ft"
    res['J1'].value = "Sites with distance from nearest tower within"+str(dq)+"miles"
    res['K1'].value = "Closest MLA"

    # read the OC sheet
    for row in range(2, oc_maxrow+1):
        l = []
        v = []
        uniq = oc['A' + str(row)].value
        lat = oc['B' + str(row)].value
        lng = oc['C' + str(row)].value
        rad = oc['D' + str(row)].value
        location1 = (lat, lng)
        print uniq,location1

        if uniq == None:
            continue
        minstuff = comparestuff(xyz,res, uniq, location1,rad, xyz_maxrow, l, v)
                # reference "cellsiteXYZ":xyuniq,"SprintCellsite":sprintcell,"location1":location1,"location2":location2,"distance":dis
        print ("processing...",uniq)
                 # print (minstuff['SprintCellsite'], minstuff['location1'], minstuff['cellsiteXYZ'], minstuff['location2'],
                 #        minstuff['distance'], minstuff['heightdiff'], '\n--------')
        res['A' + str(row)].value = minstuff['SprintCellsite']
        res['B' + str(row)].value = str(minstuff['location1'])
        res['C' + str(row)].value = minstuff['cellsiteXYZ']
        res['D' + str(row)].value = str(minstuff['location2'])
        res['E' + str(row)].value = minstuff['distance']
        res['F' + str(row)].value = minstuff['xyrad']
        res['H' + str(row)].value = minstuff['heightdiff']
        res['G' + str(row)].value = rad
        res['K' + str(row)].value = minstuff['xymla']
        vb = foreachmla(xyz,res,"VB",uniq,location1,xyz_maxrow,[])
        res['P' + str(row)].value = vb['mlauniq']
        pb = foreachmla(xyz,res,"SBA",uniq,location1,xyz_maxrow,[])
        res['O' + str(row)].value = pb['mlauniq']
        sb = foreachmla(xyz,res,"TowerCo",uniq,location1,xyz_maxrow,[])
        res['R' + str(row)].value = sb['mlauniq']
        count = 0
        for z in range(2, oc.max_row + 1):
            if (res['A' + str(z)].value) != None:
                count = count + 1
        for i in range(2, count + 2):
             if ((res['H' + str(i)].value) >= int(hq)):
                res['I' + str(i)].value = res['A' + str(i)].value
             else:
                 res['I' + str(i)].value = "Doesn't meet criteria"
        for i in range(2, count + 2):

            if ((res['E' + str(i)].value) < float(dq)):
                res['J' + str(i)].value = res['A' + str(i)].value
            else:
                res['J' + str(i)].value = "Doesn't meet criteria"


    wb.save('New6.xlsx')
def main():
    # import the worksheet
    hquery = input("Enter the query for the RAD center difference  ")
    dquery = input("Enter the query for the distance  ")
    workbook = openpyxl.load_workbook(sys.argv[1])
    doExcelProc(workbook,hquery,dquery)


if __name__ == "__main__":
    main()
