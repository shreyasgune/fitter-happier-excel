import openpyxl
import sys
from geopy.distance import vincenty


def distanceCalc(x,y):
    return vincenty(x, y).miles

def comparestuff(xyz,sprintcell,location1,maxrow,l,v):
    for xyrow in range(2,maxrow+1):
        xyuniq = xyz['A'+str(xyrow)].value
        xylat = xyz['B'+str(xyrow)].value
        xylng = xyz['C'+str(xyrow)].value
        location2 = (xylat,xylng)
        dis = distanceCalc(location1,location2)
        l.append(dis)
        v.append({"cellsiteXYZ":xyuniq,"SprintCellsite":sprintcell,"location1":location1,"location2":location2,"distance":dis})
    #print min(l)
    minIndex=l.index(min(l))
    return v[minIndex]

def doExcelProc(wb):
    print (wb.sheetnames)

    #get sheets
    oc = wb['OurCompany']
    xyz = wb['XYZcompany']
    #get sheet max and min rows and cols
    oc_maxrow = oc.max_row
    oc_maxcol = oc.max_column
    xyz_maxrow = xyz.max_row
    xyz_maxcol = xyz.max_column

    #make new Sheet
    wb.create_sheet(index=3, title='Results')
    res = wb['Results']
    res['A1'].value = "SprintCell"
    res['B1'].value = "Sprint Location"
    res['C1'].value = "XYZCell"
    res['D1'].value = "XYZCell location"
    res['E1'].value = "Minimum Distance"

    #read the OC sheet
    for row in range(2,oc_maxrow+1):
        l = []
        v = []
        uniq = oc['A'+str(row)].value
        lat = oc['B'+str(row)].value
        lng = oc['C'+str(row)].value
        location1 = (lat,lng)
        #print uniq,location1
        minstuff = comparestuff(xyz,uniq,location1,xyz_maxrow,l,v)
        #reference "cellsiteXYZ":xyuniq,"SprintCellsite":sprintcell,"location1":location1,"location2":location2,"distance":dis
        print minstuff['SprintCellsite'],minstuff['location1'],minstuff['cellsiteXYZ'],minstuff['location2'],minstuff['distance'],'\n--------'
        res['A'+str(row)].value = minstuff['SprintCellsite']
        res['B'+str(row)].value = str(minstuff['location1'])
        res['C'+str(row)].value = minstuff['cellsiteXYZ']
        res['D'+str(row)].value = str(minstuff['location2'])
        res['E'+str(row)].value = minstuff['distance']
    wb.save('ATCvsSBA.xlsx')



def main():
    #import the worksheet
    workbook = openpyxl.load_workbook(sys.argv[1])
    doExcelProc(workbook)

if __name__ == "__main__":
    main()
